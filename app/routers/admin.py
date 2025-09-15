from datetime import date, datetime, time, timedelta, timezone
from typing import List, Optional
import io
import secrets
import qrcode

from fastapi import APIRouter, Depends, Form, Request, status
from typing import Optional
from fastapi.responses import RedirectResponse, JSONResponse, StreamingResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models import User, ScheduleEntry, Store, Attendance, AllowedIP
from app.routers.attendance import _get_current_user, _check_ip_allowed, _to_moscow_time, _get_moscow_time


router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def _current_user(request: Request, db: Session) -> Optional[User]:
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    return db.get(User, user_id)


def _admin_guard(request: Request, db: Session) -> Optional[User]:
    user = _current_user(request, db)
    if not user:
        return None
    if user.role != "admin":
        return None
    return user


def _ensure_admin(request: Request, db: Session):
    user = _admin_guard(request, db)
    if not user:
        if not request.session.get("user_id"):
            return RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
        return RedirectResponse(url="/dashboard", status_code=status.HTTP_303_SEE_OTHER)
    return user


@router.get("/admin", include_in_schema=False)
def admin_root(request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    # Получаем статистику
    total_users = db.query(User).filter(User.is_active == True).count()
    total_employees = db.query(User).filter(User.role == "employee", User.is_active == True).count()
    total_admins = db.query(User).filter(User.role == "admin", User.is_active == True).count()
    total_stores = db.query(Store).count()
    today_schedules = db.query(ScheduleEntry).filter(ScheduleEntry.work_date == date.today()).count()
    
    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request, 
            "title": "Админ — панель", 
            "active_tab": "dashboard",
            "message": "Добро пожаловать в админ панель!",
            "employees": [],
            "stats": {
                "total_users": total_users,
                "total_employees": total_employees,
                "total_admins": total_admins,
                "total_stores": total_stores,
                "today_schedules": today_schedules
            }
        },
    )


@router.get("/admin/planning", include_in_schema=False)
def admin_planning(request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        # Получаем всех пользователей (сотрудников и администраторов)
        employees = db.query(User).filter(User.is_active == True).all()
        
        # Получаем все магазины
        stores = db.query(Store).all()
        
        # Получаем текущие записи расписания
        current_schedules = db.query(ScheduleEntry).filter(
            ScheduleEntry.work_date >= date.today()
        ).order_by(ScheduleEntry.work_date, ScheduleEntry.start_time).all()
        
    except Exception as e:
        print(f"Ошибка при получении данных: {e}")
        employees = []
        stores = []
        current_schedules = []
    
    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — планирование",
            "active_tab": "planning",
            "employees": employees,
            "stores": stores,
            "current_schedules": current_schedules,
            "message": "Создание и управление графиками смен",
        },
    )


@router.post("/admin/planning/create", include_in_schema=False)
def create_schedule(
    request: Request,
    employee_id: int = Form(...),
    work_date: date = Form(...),
    start_time: time = Form(...),
    end_time: time = Form(...),
    store_id: int = Form(None),
    notes: str = Form(""),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        # Проверяем, что пользователь существует
        employee = db.query(User).filter(User.id == employee_id, User.is_active == True).first()
        if not employee:
            return RedirectResponse(url="/admin/planning?error=employee_not_found", status_code=status.HTTP_303_SEE_OTHER)
        
        # Проверяем, что время корректное
        if start_time >= end_time:
            return RedirectResponse(url="/admin/planning?error=invalid_time", status_code=status.HTTP_303_SEE_OTHER)
        
        # Проверяем, что нет конфликта с существующими сменами
        existing = db.query(ScheduleEntry).filter(
            ScheduleEntry.user_id == employee_id,
            ScheduleEntry.work_date == work_date,
            ScheduleEntry.start_time < end_time,
            ScheduleEntry.end_time > start_time
        ).first()
        
        if existing:
            return RedirectResponse(url="/admin/planning?error=conflict", status_code=status.HTTP_303_SEE_OTHER)
        
        # Создаем новую запись расписания
        schedule = ScheduleEntry(
            user_id=employee_id,
            work_date=work_date,
            start_time=start_time,
            end_time=end_time,
            store_id=store_id,
            notes=notes
        )
        
        db.add(schedule)
        db.commit()
        
        return RedirectResponse(url="/admin/planning?success=created", status_code=status.HTTP_303_SEE_OTHER)
        
    except Exception as e:
        print(f"Ошибка при создании расписания: {e}")
        return RedirectResponse(url="/admin/planning?error=server_error", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/admin/planning/delete/{schedule_id}", include_in_schema=False)
def delete_schedule(schedule_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        schedule = db.query(ScheduleEntry).filter(ScheduleEntry.id == schedule_id).first()
        if schedule:
            db.delete(schedule)
            db.commit()

        return RedirectResponse(url="/admin/planning?success=deleted", status_code=status.HTTP_303_SEE_OTHER)

    except Exception as e:
        print(f"Ошибка при удалении расписания: {e}")
        return RedirectResponse(url="/admin/planning?error=delete_error", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/admin/scheduling-table", include_in_schema=False)
def admin_scheduling_table(
    request: Request,
    month: int = None,
    year: int = None,
    store_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Получаем текущий месяц если параметры не указаны
        today = date.today()
        if month is None:
            month = today.month
        if year is None:
            year = today.year

        # Валидация параметров
        if month < 1 or month > 12:
            month = today.month
        if year < 2020 or year > 2030:  # Ограничение на разумный диапазон
            year = today.year

        # Первый день выбранного месяца
        first_day = date(year, month, 1)
        # Последний день выбранного месяца
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)

        # Создаем список всех дат месяца
        month_dates = []
        current_date = first_day
        while current_date <= last_day:
            month_dates.append(current_date)
            current_date += timedelta(days=1)

        # Получаем все магазины
        stores = db.query(Store).all()

        # Получаем всех активных сотрудников и (опционально) фильтруем по магазину
        employees_query = db.query(User).filter(User.is_active == True)
        # Параметр может прийти как пустая строка — игнорируем в этом случае
        if store_id and str(store_id).strip().isdigit():
            employees_query = employees_query.filter(User.store_id == int(store_id))
        employees = employees_query.all()
        employees.sort(key=lambda u: (
            (u.store.name if getattr(u, "store", None) and u.store and u.store.name else "Без магазина").lower(),
            (u.full_name or u.email or "").lower()
        ))

        # Получаем существующие смены на этот месяц
        month_schedules = db.query(ScheduleEntry).filter(
            ScheduleEntry.work_date >= first_day,
            ScheduleEntry.work_date <= last_day
        ).all()

        # Группируем смены по сотруднику и дате
        schedule_dict = {}
        for schedule in month_schedules:
            key = f"{schedule.user_id}_{schedule.work_date}"
            schedule_dict[key] = schedule

        # Создаем записи по умолчанию только для дат, где их еще нет
        # Это позволит гибко работать с частично заполненными месяцами
        default_schedules_to_create = []
        for employee in employees:
            for work_date in month_dates:
                key = f"{employee.id}_{work_date}"
                if key not in schedule_dict:
                    # Создаем запись с рабочим днем по умолчанию (не опубликована)
                    default_schedule = ScheduleEntry(
                        user_id=employee.id,
                        work_date=work_date,
                        shift_type="work",
                        start_time=datetime.strptime('09:00', '%H:%M').time(),
                        end_time=datetime.strptime('17:00', '%H:%M').time(),
                        published=False
                    )
                    default_schedules_to_create.append(default_schedule)
                    schedule_dict[key] = default_schedule

        # Сохраняем записи по умолчанию в базу данных
        if default_schedules_to_create:
            for schedule in default_schedules_to_create:
                db.add(schedule)
            db.commit()

        # Определяем типы смен для выпадающего списка
        shift_types = [
            {"value": "work", "label": "Рабочий день"},
            {"value": "off", "label": "Отгул"},
            {"value": "weekend", "label": "Выходной"},
            {"value": "vacation", "label": "Отпуск"},
            {"value": "sick", "label": "Больничный"}
        ]

    except Exception as e:
        print(f"Ошибка при получении данных для таблицы планирования: {e}")
        month_dates = []
        employees = []
        stores = []
        schedule_dict = {}
        shift_types = []

    # Получаем список месяцев для селектора
    months = [
        {"value": 1, "name": "Январь"},
        {"value": 2, "name": "Февраль"},
        {"value": 3, "name": "Март"},
        {"value": 4, "name": "Апрель"},
        {"value": 5, "name": "Май"},
        {"value": 6, "name": "Июнь"},
        {"value": 7, "name": "Июль"},
        {"value": 8, "name": "Август"},
        {"value": 9, "name": "Сентябрь"},
        {"value": 10, "name": "Октябрь"},
        {"value": 11, "name": "Ноябрь"},
        {"value": 12, "name": "Декабрь"}
    ]

    # Получаем список лет для селектора
    current_year = date.today().year
    years = [{"value": y, "name": str(y)} for y in range(current_year - 1, current_year + 3)]

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — Таблица планирования",
            "active_tab": "scheduling_table",
            "employees": employees,
            "stores": stores,
            "month_dates": month_dates,
            "schedule_dict": schedule_dict,
            "shift_types": shift_types,
            "months": months,
            "years": years,
            "selected_month": month,
            "selected_year": year,
            "selected_store_id": (int(store_id) if store_id and str(store_id).strip().isdigit() else None),
            "message": "Интерактивная таблица планирования смен",
        },
    )


@router.post("/admin/scheduling-table/toggle", include_in_schema=False)
def toggle_schedule_slot(
    request: Request,
    employee_id: int = Form(...),
    work_date: date = Form(...),
    shift_type: str = Form(""),
    start_time: str = Form(""),
    end_time: str = Form(""),
    store_id: int = Form(None),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Проверяем, существует ли уже смена для этой даты и сотрудника
        existing = db.query(ScheduleEntry).filter(
            ScheduleEntry.user_id == employee_id,
            ScheduleEntry.work_date == work_date
        ).first()

        if shift_type == "":
            # Если выбрано пустое значение, удаляем существующую смену
            if existing:
                db.delete(existing)
                db.commit()
                return JSONResponse({"success": True, "action": "deleted"})
            else:
                return JSONResponse({"success": True, "action": "no_change"})
        else:
            # Преобразуем время в объекты time, если они указаны
            start_time_obj = None
            end_time_obj = None
            if start_time:
                try:
                    start_time_obj = datetime.strptime(start_time, '%H:%M').time()
                except:
                    pass
            if end_time:
                try:
                    end_time_obj = datetime.strptime(end_time, '%H:%M').time()
                except:
                    pass

            if existing:
                # Обновляем существующую смену
                existing.shift_type = shift_type
                existing.start_time = start_time_obj
                existing.end_time = end_time_obj
                existing.store_id = store_id if store_id else existing.store_id
                db.commit()
                return JSONResponse({"success": True, "action": "updated"})
            else:
                # Создаем новую смену
                schedule = ScheduleEntry(
                    user_id=employee_id,
                    work_date=work_date,
                    shift_type=shift_type,
                    start_time=start_time_obj,
                    end_time=end_time_obj,
                    store_id=store_id
                )
                db.add(schedule)
                db.commit()
                return JSONResponse({"success": True, "action": "created"})

    except Exception as e:
        print(f"Ошибка при изменении типа смены: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@router.post("/admin/scheduling-table/publish", include_in_schema=False)
def publish_schedules(
    request: Request,
    month: int = Form(...),
    year: int = Form(...),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Определяем диапазон дат для публикации
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)

        # Получаем все неопубликованные смены за этот месяц
        unpublished_schedules = db.query(ScheduleEntry).filter(
            ScheduleEntry.work_date >= first_day,
            ScheduleEntry.work_date <= last_day,
            ScheduleEntry.published == False
        ).all()

        # Публикуем все неопубликованные смены
        published_count = 0
        for schedule in unpublished_schedules:
            schedule.published = True
            published_count += 1

        db.commit()

        return JSONResponse({
            "success": True,
            "message": f"Опубликовано {published_count} смен за {month:02d}.{year}",
            "published_count": published_count
        })

    except Exception as e:
        print(f"Ошибка при публикации расписания: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@router.get("/admin/schedule", include_in_schema=False)
def admin_schedule(
    request: Request,
    month: int = None,
    year: int = None,
    store_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Получаем текущий месяц если параметры не указаны
        today = date.today()
        if month is None:
            month = today.month
        if year is None:
            year = today.year

        # Валидация параметров
        if month < 1 or month > 12:
            month = today.month
        if year < 2020 or year > 2030:  # Ограничение на разумный диапазон
            year = today.year

        # Первый день выбранного месяца
        first_day = date(year, month, 1)
        # Последний день выбранного месяца
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)

        # Создаем список всех дат месяца
        month_dates = []
        current_date = first_day
        while current_date <= last_day:
            month_dates.append(current_date)
            current_date += timedelta(days=1)

        # Получаем все магазины
        stores = db.query(Store).all()

        # Получаем всех активных сотрудников (опционально фильтруем по магазину) и сортируем как в таблице планирования
        employees_query = db.query(User).filter(User.is_active == True)
        if store_id and str(store_id).strip().isdigit():
            employees_query = employees_query.filter(User.store_id == int(store_id))
        employees = employees_query.all()
        employees.sort(key=lambda u: (
            (u.store.name if getattr(u, "store", None) and u.store and u.store.name else "Без магазина").lower(),
            (u.full_name or u.email or "").lower()
        ))

        # Получаем только опубликованные смены на этот месяц
        month_schedules = db.query(ScheduleEntry).filter(
            ScheduleEntry.work_date >= first_day,
            ScheduleEntry.work_date <= last_day,
            ScheduleEntry.published == True
        ).all()

        # Группируем смены по сотруднику и дате
        schedule_dict = {}
        for schedule in month_schedules:
            key = f"{schedule.user_id}_{schedule.work_date}"
            schedule_dict[key] = schedule

        # Определяем типы смен для выпадающего списка
        shift_types = [
            {"value": "work", "label": "Рабочий день"},
            {"value": "off", "label": "Отгул"},
            {"value": "weekend", "label": "Выходной"},
            {"value": "vacation", "label": "Отпуск"},
            {"value": "sick", "label": "Больничный"}
        ]

    except Exception as e:
        print(f"Ошибка при получении данных для просмотра графика: {e}")
        month_dates = []
        employees = []
        stores = []
        schedule_dict = {}
        shift_types = []

    # Получаем список месяцев для селектора
    months = [
        {"value": 1, "name": "Январь"},
        {"value": 2, "name": "Февраль"},
        {"value": 3, "name": "Март"},
        {"value": 4, "name": "Апрель"},
        {"value": 5, "name": "Май"},
        {"value": 6, "name": "Июнь"},
        {"value": 7, "name": "Июль"},
        {"value": 8, "name": "Август"},
        {"value": 9, "name": "Сентябрь"},
        {"value": 10, "name": "Октябрь"},
        {"value": 11, "name": "Ноябрь"},
        {"value": 12, "name": "Декабрь"}
    ]

    # Получаем список лет для селектора
    current_year = date.today().year
    years = [{"value": y, "name": str(y)} for y in range(current_year - 1, current_year + 3)]

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — график",
            "active_tab": "schedule",
            "employees": employees,
            "stores": stores,
            "month_dates": month_dates,
            "schedule_dict": schedule_dict,
            "shift_types": shift_types,
            "months": months,
            "years": years,
            "selected_month": month,
            "selected_year": year,
            "selected_store_id": (int(store_id) if store_id and str(store_id).strip().isdigit() else None),
            "message": "Просмотр текущих графиков и расписаний",
        },
    )


@router.get("/admin/reports", include_in_schema=False)
def admin_reports(
    request: Request,
    report_type: str = "month",
    start_date: str = None,
    end_date: str = None,
    month: int = None,
    year: int = None,
    store_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Определяем даты для отчета
        today = date.today()

        if report_type == "month":
            # Выбранный месяц или текущий месяц
            selected_month = month if month is not None else today.month
            selected_year = year if year is not None else today.year

            # Валидация параметров
            if selected_month < 1 or selected_month > 12:
                selected_month = today.month
            if selected_year < 2020 or selected_year > 2030:
                selected_year = today.year

            start_date_obj = date(selected_year, selected_month, 1)
            if selected_month == 12:
                end_date_obj = date(selected_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date_obj = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
        elif report_type == "year":
            # Текущий год
            start_date_obj = date(today.year, 1, 1)
            end_date_obj = date(today.year, 12, 31)
        elif report_type == "custom" and start_date and end_date:
            # Произвольный интервал
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            except:
                # Если ошибка в датах, используем текущий месяц
                start_date_obj = date(today.year, today.month, 1)
                if today.month == 12:
                    end_date_obj = date(today.year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date_obj = date(today.year, today.month + 1, 1) - timedelta(days=1)
        else:
            # По умолчанию текущий месяц
            start_date_obj = date(today.year, today.month, 1)
            if today.month == 12:
                end_date_obj = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date_obj = date(today.year, today.month + 1, 1) - timedelta(days=1)

        # Получаем всех активных сотрудников (опционально по магазину) и сортируем по магазину, затем по имени
        employees_query = db.query(User).filter(User.is_active == True)
        if store_id and str(store_id).strip().isdigit():
            employees_query = employees_query.filter(User.store_id == int(store_id))
        employees = employees_query.all()
        employees.sort(key=lambda u: (
            (u.store.name if getattr(u, "store", None) and u.store and u.store.name else "Без магазина").lower(),
            (u.full_name or u.email or "").lower()
        ))

        # Собираем данные для каждого сотрудника
        report_data = []
        for employee in employees:
            # Получаем записи посещаемости за период
            attendances = db.query(Attendance).filter(
                Attendance.user_id == employee.id,
                Attendance.work_date >= start_date_obj,
                Attendance.work_date <= end_date_obj
            ).all()

            # Получаем записи расписания за период
            schedules = db.query(ScheduleEntry).filter(
                ScheduleEntry.user_id == employee.id,
                ScheduleEntry.work_date >= start_date_obj,
                ScheduleEntry.work_date <= end_date_obj,
                ScheduleEntry.published == True
            ).all()

            # Рассчитываем статистику - правильно подсчитываем время за каждый день
            total_hours = 0
            working_shifts = 0
            
            # Группируем записи по дням
            daily_attendance = {}
            for att in attendances:
                work_date = att.work_date
                if work_date not in daily_attendance:
                    daily_attendance[work_date] = []
                daily_attendance[work_date].append(att)
            
            # Подсчитываем время для каждого дня
            for work_date, day_attendances in daily_attendance.items():
                # Сортируем записи по времени начала
                day_attendances.sort(key=lambda x: x.started_at)
                
                # Подсчитываем общее время за день
                day_total_seconds = 0
                day_has_work = False
                
                for att in day_attendances:
                    if att.ended_at is not None:
                        # Завершенная сессия - добавляем её время
                        started_at = att.started_at
                        ended_at = att.ended_at
                        
                        # Convert to Moscow timezone for consistent calculations
                        started_at = _to_moscow_time(started_at)
                        ended_at = _to_moscow_time(ended_at)

                        session_seconds = (ended_at - started_at).total_seconds()
                        day_total_seconds += session_seconds
                        day_has_work = True
                
                if day_has_work:
                    day_hours = day_total_seconds / 3600.0
                    total_hours += day_hours
                    working_shifts += 1

            # Считаем типы смен из расписания
            days_off = len([s for s in schedules if s.shift_type == 'off'])
            vacations = len([s for s in schedules if s.shift_type == 'vacation'])
            sick_days = len([s for s in schedules if s.shift_type == 'sick'])
            work_days = len([s for s in schedules if s.shift_type == 'work'])

            report_data.append({
                'employee': employee,
                'total_hours': round(total_hours, 2),
                'working_shifts': working_shifts,
                'work_days': work_days,
                'days_off': days_off,
                'vacations': vacations,
                'sick_days': sick_days,
                'attendances': attendances,
                'schedules': schedules
            })

        # Общая статистика
        total_employees = len(employees)
        total_hours_all = sum(data['total_hours'] for data in report_data)
        total_shifts_all = sum(data['working_shifts'] for data in report_data)

        # Получаем списки месяцев и лет для селекторов
        months = [
            {"value": 1, "name": "Январь"},
            {"value": 2, "name": "Февраль"},
            {"value": 3, "name": "Март"},
            {"value": 4, "name": "Апрель"},
            {"value": 5, "name": "Май"},
            {"value": 6, "name": "Июнь"},
            {"value": 7, "name": "Июль"},
            {"value": 8, "name": "Август"},
            {"value": 9, "name": "Сентябрь"},
            {"value": 10, "name": "Октябрь"},
            {"value": 11, "name": "Ноябрь"},
            {"value": 12, "name": "Декабрь"}
        ]

        current_year = date.today().year
        years = [{"value": y, "name": str(y)} for y in range(current_year - 1, current_year + 3)]

        # Определяем выбранные значения
        selected_month = month if month is not None else today.month
        selected_year = year if year is not None else today.year

    except Exception as e:
        print(f"Ошибка при получении отчетов: {e}")
        report_data = []
        total_employees = 0
        total_hours_all = 0
        total_shifts_all = 0
        start_date_obj = date.today()
        end_date_obj = date.today()

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — отчеты",
            "active_tab": "reports",
            "report_data": report_data,
            "report_type": report_type,
            "start_date": start_date_obj.strftime('%Y-%m-%d') if start_date_obj else None,
            "end_date": end_date_obj.strftime('%Y-%m-%d') if end_date_obj else None,
            "total_employees": total_employees,
            "total_hours_all": round(total_hours_all, 2),
            "total_shifts_all": total_shifts_all,
            "months": months,
            "years": years,
            "selected_month": selected_month,
            "selected_year": selected_year,
            "stores": db.query(Store).all(),
            "selected_store_id": (int(store_id) if store_id and str(store_id).strip().isdigit() else None),
            "message": "Аналитика и отчеты по рабочему времени",
        },
    )


@router.get("/admin/reports/export", include_in_schema=False)
def export_reports(
    request: Request,
    report_type: str = "month",
    start_date: str = None,
    end_date: str = None,
    month: int = None,
    year: int = None,
    store_id: int = None,
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Определяем даты для отчета (повторяем логику из основного отчета)
        today = date.today()

        if report_type == "month":
            # Выбранный месяц или текущий месяц
            selected_month = month if month is not None else today.month
            selected_year = year if year is not None else today.year

            # Валидация параметров
            if selected_month < 1 or selected_month > 12:
                selected_month = today.month
            if selected_year < 2020 or selected_year > 2030:
                selected_year = today.year

            start_date_obj = date(selected_year, selected_month, 1)
            if selected_month == 12:
                end_date_obj = date(selected_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date_obj = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
        elif report_type == "year":
            start_date_obj = date(today.year, 1, 1)
            end_date_obj = date(today.year, 12, 31)
        elif report_type == "custom" and start_date and end_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            except:
                start_date_obj = date(today.year, today.month, 1)
                if today.month == 12:
                    end_date_obj = date(today.year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date_obj = date(today.year, today.month + 1, 1) - timedelta(days=1)
        else:
            start_date_obj = date(today.year, today.month, 1)
            if today.month == 12:
                end_date_obj = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date_obj = date(today.year, today.month + 1, 1) - timedelta(days=1)

        # Получаем всех активных сотрудников (с фильтром по магазину, если указан)
        employees_query = db.query(User).filter(User.is_active == True)
        if store_id:
            employees_query = employees_query.filter(User.store_id == store_id)
        employees = employees_query.order_by(User.full_name).all()

        # Собираем данные для каждого сотрудника
        report_data = []
        for employee in employees:
            attendances = db.query(Attendance).filter(
                Attendance.user_id == employee.id,
                Attendance.work_date >= start_date_obj,
                Attendance.work_date <= end_date_obj
            ).all()

            schedules = db.query(ScheduleEntry).filter(
                ScheduleEntry.user_id == employee.id,
                ScheduleEntry.work_date >= start_date_obj,
                ScheduleEntry.work_date <= end_date_obj,
                ScheduleEntry.published == True
            ).all()

            # Рассчитываем статистику - правильно подсчитываем время за каждый день
            total_hours = 0
            working_shifts = 0
            
            # Группируем записи по дням
            daily_attendance = {}
            for att in attendances:
                work_date = att.work_date
                if work_date not in daily_attendance:
                    daily_attendance[work_date] = []
                daily_attendance[work_date].append(att)
            
            # Подсчитываем время для каждого дня
            for work_date, day_attendances in daily_attendance.items():
                # Сортируем записи по времени начала
                day_attendances.sort(key=lambda x: x.started_at)
                
                # Подсчитываем общее время за день
                day_total_seconds = 0
                day_has_work = False
                
                for att in day_attendances:
                    if att.ended_at is not None:
                        # Завершенная сессия - добавляем её время
                        started_at = att.started_at
                        ended_at = att.ended_at
                        
                        # Convert to Moscow timezone for consistent calculations
                        started_at = _to_moscow_time(started_at)
                        ended_at = _to_moscow_time(ended_at)

                        session_seconds = (ended_at - started_at).total_seconds()
                        day_total_seconds += session_seconds
                        day_has_work = True
                
                if day_has_work:
                    day_hours = day_total_seconds / 3600.0
                    total_hours += day_hours
                    working_shifts += 1

            days_off = len([s for s in schedules if s.shift_type == 'off'])
            vacations = len([s for s in schedules if s.shift_type == 'vacation'])
            sick_days = len([s for s in schedules if s.shift_type == 'sick'])
            work_days = len([s for s in schedules if s.shift_type == 'work'])

            report_data.append({
                'employee': employee,
                'total_hours': round(total_hours, 2),
                'working_shifts': working_shifts,
                'work_days': work_days,
                'days_off': days_off,
                'vacations': vacations,
                'sick_days': sick_days,
            })

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по рабочему времени"

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        total_font = Font(bold=True, color="000000")
        total_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            "Сотрудник", "Email", "Рабочие часы", "Рабочие смены",
            "Рабочие дни (по графику)", "Отгулы", "Отпуска", "Больничные",
            "Среднее время за смену"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Группировка по магазинам в Excel
        # Подготовка стилей для заголовков групп (зелёные, как на сайте)
        store_header_font = Font(bold=True, color="155724")
        store_header_fill = PatternFill(start_color="EAF7EA", end_color="EAF7EA", fill_type="solid")
        store_header_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'), bottom=Side(style='thin')
        )

        # Сортируем данные как в HTML: по названию магазина (или "Без магазина"), затем по имени/почте
        def _store_name(d):
            emp = d['employee']
            return (emp.store.name if getattr(emp, 'store', None) else 'Без магазина')

        def _employee_key(d):
            emp = d['employee']
            return (emp.full_name or emp.email or '').lower()

        report_data_sorted = sorted(report_data, key=lambda d: (_store_name(d) or '', _employee_key(d)))

        current_row = 2
        current_store = None
        for data in report_data_sorted:
            employee = data['employee']
            emp_store_name = employee.store.name if getattr(employee, 'store', None) else 'Без магазина'

            # При смене магазина — вставляем заголовок группы
            if current_store != emp_store_name:
                current_store = emp_store_name
                # Заголовок на всю ширину таблицы
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                header_cell = ws.cell(row=current_row, column=1, value=f"🏪 {current_store}")
                header_cell.font = store_header_font
                header_cell.fill = store_header_fill
                header_cell.alignment = Alignment(horizontal="left")
                # Рамка по всей строке заголовка
                for col_num in range(1, len(headers) + 1):
                    c = ws.cell(row=current_row, column=col_num)
                    c.border = store_header_border
                current_row += 1

            # Строка сотрудника
            row_data = [
                employee.full_name or employee.email,
                employee.email,
                data['total_hours'],
                data['working_shifts'],
                data['work_days'],
                data['days_off'],
                data['vacations'],
                data['sick_days'],
                round(data['total_hours'] / data['working_shifts'], 1) if data['working_shifts'] > 0 else 0
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = border
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="center")
            current_row += 1

        # Итоговая строка
        total_row = current_row + 1
        total_hours_all = sum(data['total_hours'] for data in report_data)
        total_shifts_all = sum(data['working_shifts'] for data in report_data)

        ws.cell(row=total_row, column=1, value="ИТОГО").font = total_font
        ws.cell(row=total_row, column=3, value=total_hours_all).font = total_font
        ws.cell(row=total_row, column=4, value=total_shifts_all).font = total_font

        # Применяем границы к итоговой строке
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=col_num)
            cell.border = border
            cell.fill = total_fill
            if col_num == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)  # Максимальная ширина 30
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем в memory buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Формируем имя файла (только ASCII символы для совместимости)
        period_name = ""
        if report_type == "month":
            period_name = f"{start_date_obj.strftime('%m.%Y')}"
        elif report_type == "year":
            period_name = f"{start_date_obj.year}"
        else:
            period_name = f"{start_date_obj.strftime('%d.%m.%Y')}-{end_date_obj.strftime('%d.%m.%Y')}"

        # Используем ASCII название файла
        filename = f"work_time_report_{period_name}.xlsx"

        # Отдаем как обычный ответ с байтами файла — это надежнее для большинства браузеров
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Ошибка при экспорте в Excel: {e}")
        return RedirectResponse(
            url="/admin/reports?error=export_failed",
            status_code=status.HTTP_303_SEE_OTHER
        )


@router.get("/admin/attendance", include_in_schema=False)
def admin_attendance(
    request: Request,
    employee_id: int = None,
    selected_date: str = None,
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Получаем всех активных сотрудников
        employees = db.query(User).filter(User.is_active == True).order_by(User.full_name).all()

        attendance_record = None
        if employee_id and selected_date:
            try:
                work_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
                attendance_record = db.query(Attendance).filter(
                    Attendance.user_id == employee_id,
                    Attendance.work_date == work_date
                ).order_by(Attendance.started_at.desc()).first()
            except:
                pass

        # Получаем последние записи присутствия для быстрого доступа
        recent_attendances = db.query(Attendance).join(User).filter(
            Attendance.work_date >= date.today() - timedelta(days=7)
        ).order_by(Attendance.work_date.desc(), Attendance.started_at.desc()).limit(10).all()

    except Exception as e:
        print(f"Ошибка при получении данных о присутствии: {e}")
        employees = []
        attendance_record = None
        recent_attendances = []

    # Convert attendance times to Moscow timezone for display
    moscow_attendance_record = None
    if attendance_record:
        moscow_attendance_record = type('AttendanceDisplay', (), {
            'id': attendance_record.id,
            'user_id': attendance_record.user_id,
            'work_date': attendance_record.work_date,
            'started_at': _to_moscow_time(attendance_record.started_at) if attendance_record.started_at else None,
            'ended_at': _to_moscow_time(attendance_record.ended_at) if attendance_record.ended_at else None,
            'hours': attendance_record.hours
        })()
    
    moscow_recent_attendances = []
    for attendance in recent_attendances:
        moscow_att = type('AttendanceDisplay', (), {
            'id': attendance.id,
            'user_id': attendance.user_id,
            'work_date': attendance.work_date,
            'started_at': _to_moscow_time(attendance.started_at) if attendance.started_at else None,
            'ended_at': _to_moscow_time(attendance.ended_at) if attendance.ended_at else None,
            'hours': attendance.hours,
            'user': attendance.user
        })()
        moscow_recent_attendances.append(moscow_att)

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — Редактирование присутствия",
            "active_tab": "attendance",
            "employees": employees,
            "attendance_record": moscow_attendance_record,
            "selected_employee_id": employee_id,
            "selected_date": selected_date,
            "recent_attendances": moscow_recent_attendances,
            "message": "Редактирование записей о приходе и уходе сотрудников",
        },
    )


@router.post("/admin/attendance/save", include_in_schema=False)
def save_attendance(
    request: Request,
    employee_id: int = Form(...),
    work_date: date = Form(...),
    start_time: str = Form(""),
    end_time: str = Form(""),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Проверяем, существует ли уже запись
        existing = db.query(Attendance).filter(
            Attendance.user_id == employee_id,
            Attendance.work_date == work_date
        ).first()

        # Преобразуем время
        start_time_obj = None
        end_time_obj = None

        if start_time:
            try:
                start_time_obj = datetime.strptime(start_time, '%H:%M').time()
            except:
                pass

        if end_time:
            try:
                end_time_obj = datetime.strptime(end_time, '%H:%M').time()
            except:
                pass

        if existing:
            # Обновляем существующую запись
            if start_time_obj:
                existing.started_at = _get_moscow_time().replace(hour=start_time_obj.hour, minute=start_time_obj.minute, second=0, microsecond=0)
            if end_time_obj:
                existing.ended_at = _get_moscow_time().replace(hour=end_time_obj.hour, minute=end_time_obj.minute, second=0, microsecond=0)

                # Пересчитываем часы
                if existing.started_at and existing.ended_at:
                    elapsed_seconds = (existing.ended_at - existing.started_at).total_seconds()
                    existing.hours = round(elapsed_seconds / 3600.0, 4)
                else:
                    existing.hours = None

            db.commit()
            return RedirectResponse(
                url=f"/admin/attendance?employee_id={employee_id}&selected_date={work_date}&success=updated",
                status_code=status.HTTP_303_SEE_OTHER
            )
        else:
            # Создаем новую запись
            if start_time_obj:
                started_at = _get_moscow_time().replace(hour=start_time_obj.hour, minute=start_time_obj.minute, second=0, microsecond=0)
                attendance = Attendance(
                    user_id=employee_id,
                    started_at=started_at,
                    work_date=work_date
                )

                if end_time_obj:
                    attendance.ended_at = _get_moscow_time().replace(hour=end_time_obj.hour, minute=end_time_obj.minute, second=0, microsecond=0)
                    elapsed_seconds = (attendance.ended_at - attendance.started_at).total_seconds()
                    attendance.hours = round(elapsed_seconds / 3600.0, 4)

                db.add(attendance)
                db.commit()

                return RedirectResponse(
                    url=f"/admin/attendance?employee_id={employee_id}&selected_date={work_date}&success=created",
                    status_code=status.HTTP_303_SEE_OTHER
                )
            else:
                return RedirectResponse(
                    url=f"/admin/attendance?employee_id={employee_id}&selected_date={work_date}&error=no_start_time",
                    status_code=status.HTTP_303_SEE_OTHER
                )

    except Exception as e:
        print(f"Ошибка при сохранении присутствия: {e}")
        return RedirectResponse(
            url=f"/admin/attendance?employee_id={employee_id}&selected_date={work_date}&error=server_error",
            status_code=status.HTTP_303_SEE_OTHER
        )


@router.post("/admin/attendance/delete/{attendance_id}", include_in_schema=False)
def delete_attendance(attendance_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        attendance = db.query(Attendance).filter(Attendance.id == attendance_id).first()
        if not attendance:
            return RedirectResponse(
                url="/admin/attendance?error=not_found",
                status_code=status.HTTP_303_SEE_OTHER
            )

        employee_id = attendance.user_id
        work_date = attendance.work_date

        try:
            # Удаляем все записи за этот день для пользователя, чтобы не оставались дубликаты
            duplicates = db.query(Attendance).filter(
                Attendance.user_id == employee_id,
                Attendance.work_date == work_date
            ).all()
            for rec in duplicates:
                db.delete(rec)
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"Ошибка при удалении присутствия (id={attendance_id}): {e}")
            return RedirectResponse(
                url=f"/admin/attendance?employee_id={employee_id}&selected_date={work_date}&error=delete_error",
                status_code=status.HTTP_303_SEE_OTHER
            )

        return RedirectResponse(
            url=f"/admin/attendance?employee_id={employee_id}&selected_date={work_date}&success=deleted",
            status_code=status.HTTP_303_SEE_OTHER
        )

    except Exception as e:
        print(f"Ошибка при удалении присутствия (общая): {e}")
        return RedirectResponse(
            url="/admin/attendance?error=delete_error",
            status_code=status.HTTP_303_SEE_OTHER
        )


@router.get("/admin/attendance/delete/{attendance_id}", include_in_schema=False)
def delete_attendance_get(attendance_id: int, request: Request, db: Session = Depends(get_db)):
    """Зеркальный GET-эндпоинт удаления на случай, если форма/ссылка вызывает GET."""
    return delete_attendance(attendance_id, request, db)


@router.get("/admin/stores", include_in_schema=False)
def admin_stores(request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        # Получаем все магазины
        stores = db.query(Store).all()
        
    except Exception as e:
        print(f"Ошибка при получении магазинов: {e}")
        stores = []
    
    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — магазины",
            "active_tab": "stores",
            "stores": stores,
            "message": "Управление магазинами и филиалами",
        },
    )


@router.post("/admin/stores/{store_id}/qr-regenerate", include_in_schema=False)
def regenerate_store_qr(store_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    store = db.get(Store, store_id)
    if not store:
        return RedirectResponse(url="/admin/stores?error=store_not_found", status_code=status.HTTP_303_SEE_OTHER)

    store.qr_token = secrets.token_urlsafe(24)
    db.add(store)
    db.commit()
    return RedirectResponse(url=f"/admin/stores?ok=qr_updated", status_code=status.HTTP_303_SEE_OTHER)


@router.get("/admin/stores/{store_id}/qr.png", include_in_schema=False)
def store_qr_image(store_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    store = db.get(Store, store_id)
    if not store or not store.qr_token:
        return RedirectResponse(url="/admin/stores?error=no_qr", status_code=status.HTTP_303_SEE_OTHER)

    base_url = str(request.base_url).rstrip('/')
    qr_url = f"{base_url}/q/start/{store.qr_token}"

    img = qrcode.make(qr_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")


@router.get("/admin/stores/{store_id}/qr-stop.png", include_in_schema=False)
def store_qr_stop_image(store_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    store = db.get(Store, store_id)
    if not store or not store.qr_token:
        return RedirectResponse(url="/admin/stores?error=no_qr", status_code=status.HTTP_303_SEE_OTHER)

    base_url = str(request.base_url).rstrip('/')
    qr_url = f"{base_url}/q/stop/{store.qr_token}"

    img = qrcode.make(qr_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")


@router.get("/admin/employees", include_in_schema=False)
def admin_employees(request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        # Получаем всех пользователей
        employees = db.query(User).filter(User.is_active == True).all()
        
        # Получаем все магазины
        stores = db.query(Store).all()
        
    except Exception as e:
        print(f"Ошибка при получении данных: {e}")
        employees = []
        stores = []
    
    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — сотрудники",
            "active_tab": "employees",
            "employees": employees,
            "stores": stores,
            "message": "Управление сотрудниками и назначение магазинов",
        },
    )


@router.post("/admin/employees/assign-store", include_in_schema=False)
def assign_store_to_employee(
    request: Request,
    employee_id: int = Form(...),
    store_id: int = Form(None),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        # Находим сотрудника
        employee = db.query(User).filter(User.id == employee_id).first()
        if not employee:
            return RedirectResponse(url="/admin/employees?error=employee_not_found", status_code=status.HTTP_303_SEE_OTHER)
        
        # Назначаем магазин (может быть None для снятия назначения)
        employee.store_id = store_id if store_id else None
        db.commit()
        
        return RedirectResponse(url="/admin/employees?success=store_assigned", status_code=status.HTTP_303_SEE_OTHER)
        
    except Exception as e:
        print(f"Ошибка при назначении магазина: {e}")
        return RedirectResponse(url="/admin/employees?error=server_error", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/admin/employees/toggle-status", include_in_schema=False)
def toggle_employee_status(
    request: Request,
    employee_id: int = Form(...),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        # Находим сотрудника
        employee = db.query(User).filter(User.id == employee_id).first()
        if not employee:
            return RedirectResponse(url="/admin/employees?error=employee_not_found", status_code=status.HTTP_303_SEE_OTHER)
        
        # Переключаем статус
        employee.is_active = not employee.is_active
        db.commit()
        
        return RedirectResponse(url="/admin/employees?success=status_changed", status_code=status.HTTP_303_SEE_OTHER)
        
    except Exception as e:
        print(f"Ошибка при изменении статуса: {e}")
        return RedirectResponse(url="/admin/employees?error=server_error", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/admin/stores/create", include_in_schema=False)
def create_store(
    request: Request,
    name: str = Form(...),
    address: str = Form(""),
    phone: str = Form(""),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result
    
    try:
        store = Store(
            name=name,
            address=address,
            phone=phone
        )
        
        db.add(store)
        db.commit()
        
        return RedirectResponse(url="/admin/stores?success=created", status_code=status.HTTP_303_SEE_OTHER)
        
    except Exception as e:
        print(f"Ошибка при создании магазина: {e}")
@router.get("/admin/allowed-ips", include_in_schema=False)
def admin_allowed_ips(request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Получаем все разрешенные IP
        allowed_ips = db.query(AllowedIP).order_by(AllowedIP.created_at.desc()).all()

    except Exception as e:
        print(f"Ошибка при получении разрешенных IP: {e}")
        allowed_ips = []

    return templates.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "title": "Админ — Разрешенные IP",
            "active_tab": "allowed_ips",
            "allowed_ips": allowed_ips,
            "message": "Управление списком разрешенных IP адресов",
        },
    )


@router.post("/admin/allowed-ips/create", include_in_schema=False)
def create_allowed_ip(
    request: Request,
    ip_address: str = Form(...),
    description: str = Form(""),
    db: Session = Depends(get_db)
):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        # Проверяем, что IP не существует
        existing = db.query(AllowedIP).filter(AllowedIP.ip_address == ip_address).first()
        if existing:
            return RedirectResponse(url="/admin/allowed-ips?error=ip_exists", status_code=status.HTTP_303_SEE_OTHER)

        # Получаем текущего пользователя
        admin_user = _current_user(request, db)

        # Создаем новый разрешенный IP
        allowed_ip = AllowedIP(
            ip_address=ip_address,
            description=description,
            created_by=admin_user.id if admin_user else None
        )

        db.add(allowed_ip)
        db.commit()

        return RedirectResponse(url="/admin/allowed-ips?success=created", status_code=status.HTTP_303_SEE_OTHER)

    except Exception as e:
        print(f"Ошибка при создании разрешенного IP: {e}")
        return RedirectResponse(url="/admin/allowed-ips?error=server_error", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/admin/allowed-ips/delete/{ip_id}", include_in_schema=False)
def delete_allowed_ip(ip_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        allowed_ip = db.query(AllowedIP).filter(AllowedIP.id == ip_id).first()
        if allowed_ip:
            db.delete(allowed_ip)
            db.commit()

        return RedirectResponse(url="/admin/allowed-ips?success=deleted", status_code=status.HTTP_303_SEE_OTHER)

    except Exception as e:
        print(f"Ошибка при удалении разрешенного IP: {e}")
        return RedirectResponse(url="/admin/allowed-ips?error=delete_error", status_code=status.HTTP_303_SEE_OTHER)


@router.post("/admin/allowed-ips/toggle/{ip_id}", include_in_schema=False)
def toggle_allowed_ip_status(ip_id: int, request: Request, db: Session = Depends(get_db)):
    result = _ensure_admin(request, db)
    if isinstance(result, RedirectResponse):
        return result

    try:
        allowed_ip = db.query(AllowedIP).filter(AllowedIP.id == ip_id).first()
        if allowed_ip:
            allowed_ip.is_active = not allowed_ip.is_active
            db.commit()

        return RedirectResponse(url="/admin/allowed-ips?success=status_changed", status_code=status.HTTP_303_SEE_OTHER)

    except Exception as e:
        print(f"Ошибка при изменении статуса IP: {e}")
        return RedirectResponse(url="/admin/allowed-ips?error=status_error", status_code=status.HTTP_303_SEE_OTHER)
        return RedirectResponse(url="/admin/stores?error=server_error", status_code=status.HTTP_303_SEE_OTHER)
